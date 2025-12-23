"""
Excel Exporter

Exports compensation plans to professionally formatted Excel workbooks.
Mac & PC compatible using openpyxl.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Optional
from datetime import datetime

try:
    from ..core.library_manager import LibraryManager
    from ..core.template_engine import Plan
except ImportError:
    # When run as script
    import sys
    from pathlib import Path
    sys.path.insert(0, str(Path(__file__).parent.parent))
    from core.library_manager import LibraryManager
    from core.template_engine import Plan


class ExcelExporter:
    """
    Export compensation plans to Excel workbooks with professional formatting.
    """

    def __init__(self, library: Optional[LibraryManager] = None):
        """
        Initialize exporter.

        Args:
            library: LibraryManager instance. If None, creates new one.
        """
        self.library = library or LibraryManager()

        # Define styles
        self.title_font = Font(bold=True, size=18, color="1F4E78")
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.section_font = Font(bold=True, size=14, color="1F4E78")
        self.subsection_font = Font(bold=True, size=11, color="44546A")
        self.body_font = Font(size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def export(self, plan: Plan, filepath: Path, include_toc: bool = True) -> Path:
        """
        Export plan to Excel workbook.

        Args:
            plan: Plan to export
            filepath: Output file path
            include_toc: Include table of contents

        Returns:
            Path to created file
        """
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create sheets
        self._create_summary_sheet(wb, plan)
        self._create_full_policy_sheet(wb, plan)
        self._create_components_sheet(wb, plan)
        self._create_governance_sheet(wb, plan)

        # Save
        wb.save(filepath)

        return filepath

    def _create_summary_sheet(self, wb: openpyxl.Workbook, plan: Plan) -> None:
        """Create summary/overview sheet."""
        ws = wb.create_sheet("Summary", 0)

        row = 1

        # Title
        ws[f'A{row}'] = plan.metadata.name.upper()
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 2

        # Metadata table
        metadata_fields = [
            ("Plan ID:", plan.plan_id),
            ("Effective Date:", plan.metadata.effective_date),
            ("End Date:", plan.metadata.end_date or "N/A"),
            ("Status:", plan.metadata.status.upper()),
            ("Version:", plan.metadata.version),
            ("Author:", plan.metadata.author),
            ("Created:", plan.metadata.created_date.split('T')[0]),
            ("Last Modified:", plan.metadata.last_modified.split('T')[0]),
        ]

        for label, value in metadata_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1

        row += 1

        # Description
        if plan.metadata.description:
            ws[f'A{row}'] = "Description:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = plan.metadata.description
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            ws.merge_cells(f'A{row}:D{row}')
            row += 2

        # Components summary
        ws[f'A{row}'] = "PLAN STRUCTURE"
        ws[f'A{row}'].font = self.section_font
        row += 1

        # Section count
        ws[f'A{row}'] = f"Total Sections: {len(plan.sections)}"
        ws[f'C{row}'] = f"Total Components: {len(plan.get_all_component_ids())}"
        row += 2

        # Sections table header
        ws[f'A{row}'] = "Section"
        ws[f'B{row}'] = "Title"
        ws[f'C{row}'] = "Components"
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{row}']
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        row += 1

        # Sections
        for section in plan.sections:
            ws[f'A{row}'] = section.section_num
            ws[f'B{row}'] = section.title
            ws[f'C{row}'] = len(section.component_ids)
            row += 1

        # Format columns
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30

    def _create_full_policy_sheet(self, wb: openpyxl.Workbook, plan: Plan) -> None:
        """Create full policy document sheet."""
        ws = wb.create_sheet("Full Policy", 1)

        row = 1

        # Title
        ws[f'A{row}'] = plan.metadata.name.upper()
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 2

        ws[f'A{row}'] = f"Effective: {plan.metadata.effective_date}"
        ws[f'A{row}'].font = Font(italic=True, size=10)
        row += 2

        # Table of contents
        ws[f'A{row}'] = "TABLE OF CONTENTS"
        ws[f'A{row}'].font = self.section_font
        row += 1

        toc_start = row
        for section in plan.sections:
            ws[f'A{row}'] = f"{section.section_num}. {section.title}"
            ws[f'A{row}'].font = Font(underline="single", color="0563C1")
            row += 1
        row += 1

        # Sections
        for section in plan.sections:
            # Section header
            ws[f'A{row}'] = f"{section.section_num}. {section.title.upper()}"
            ws[f'A{row}'].font = self.section_font
            ws.merge_cells(f'A{row}:D{row}')
            row += 2

            # Components in this section
            for comp_id in section.component_ids:
                component = self.library.get_component(comp_id)

                # Component name
                ws[f'A{row}'] = component.name
                ws[f'A{row}'].font = self.subsection_font
                ws.merge_cells(f'A{row}:D{row}')
                row += 1

                # Policy text (detailed)
                policy_cell = ws[f'A{row}']
                policy_cell.value = component.policy_text.detailed
                policy_cell.alignment = Alignment(wrap_text=True, vertical='top')
                ws.merge_cells(f'A{row}:D{row}')
                ws.row_dimensions[row].height = None  # Auto height
                row += 1

                # Legal text if present
                if component.policy_text.legal:
                    legal_cell = ws[f'A{row}']
                    legal_cell.value = f"Legal: {component.policy_text.legal}"
                    legal_cell.font = Font(italic=True, size=9, color="7F7F7F")
                    legal_cell.alignment = Alignment(wrap_text=True, vertical='top')
                    ws.merge_cells(f'A{row}:D{row}')
                    row += 1

                row += 1  # Space between components

            # Custom text for section if present
            if section.custom_text:
                ws[f'A{row}'] = section.custom_text
                ws[f'A{row}'].alignment = Alignment(wrap_text=True)
                ws.merge_cells(f'A{row}:D{row}')
                row += 2

            row += 1  # Space between sections

        # Format columns
        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

    def _create_components_sheet(self, wb: openpyxl.Workbook, plan: Plan) -> None:
        """Create components detail sheet."""
        ws = wb.create_sheet("Components", 2)

        row = 1

        # Title
        ws[f'A{row}'] = "COMPONENT DETAILS"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:F{row}')
        row += 2

        # Headers
        headers = ['ID', 'Name', 'Category', 'Version', 'Requirements', 'Conflicts']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        row += 1

        # Components
        for comp_id in sorted(plan.get_all_component_ids()):
            component = self.library.get_component(comp_id)

            ws.cell(row=row, column=1).value = component.component_id
            ws.cell(row=row, column=2).value = component.name
            ws.cell(row=row, column=3).value = component.category
            ws.cell(row=row, column=4).value = component.version
            ws.cell(row=row, column=5).value = ", ".join(component.validation_rules.requires) if component.validation_rules.requires else "-"
            ws.cell(row=row, column=6).value = ", ".join(component.validation_rules.conflicts_with) if component.validation_rules.conflicts_with else "-"

            row += 1

        # Format columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 25

    def _create_governance_sheet(self, wb: openpyxl.Workbook, plan: Plan) -> None:
        """Create governance scorecard sheet."""
        ws = wb.create_sheet("Governance", 3)

        row = 1

        # Title
        ws[f'A{row}'] = "GOVERNANCE SCORECARD"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:E{row}')
        row += 2

        # Calculate scores
        total_financial = 0
        total_complexity = 0
        total_compliance = 0
        total_operational = 0
        num_components = len(plan.get_all_component_ids())

        for comp_id in plan.get_all_component_ids():
            component = self.library.get_component(comp_id)
            total_financial += component.governance_impact.financial_risk
            total_complexity += component.governance_impact.complexity
            total_compliance += component.governance_impact.compliance_risk
            total_operational += component.governance_impact.operational_risk

        # Calculate weighted score (out of 100)
        financial_score = min(100, (total_financial / max(1, num_components)) * 10 * 0.30)
        complexity_score = min(100, (total_complexity / max(1, num_components)) * 10 * 0.25)
        compliance_score = min(100, (total_compliance / max(1, num_components)) * 10 * 0.25)
        operational_score = min(100, (total_operational / max(1, num_components)) * 10 * 0.20)
        total_score = int(financial_score + complexity_score + compliance_score + operational_score)

        # Risk level
        if total_score < 40:
            risk_level = "LOW RISK"
            risk_color = "00B050"
        elif total_score < 70:
            risk_level = "MEDIUM RISK"
            risk_color = "FFC000"
        else:
            risk_level = "HIGH RISK"
            risk_color = "C00000"

        # Overall score
        ws[f'A{row}'] = "Overall Governance Risk Score:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'] = f"{total_score}/100"
        ws[f'B{row}'].font = Font(bold=True, size=14, color=risk_color)
        row += 1

        ws[f'A{row}'] = "Risk Level:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = risk_level
        ws[f'B{row}'].font = Font(bold=True, color=risk_color)
        row += 2

        # Breakdown table
        ws[f'A{row}'] = "RISK BREAKDOWN"
        ws[f'A{row}'].font = self.section_font
        row += 1

        # Headers
        headers = ['Category', 'Weight', 'Raw Score', 'Weighted Score', 'Max']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        row += 1

        # Rows
        breakdown = [
            ("Financial Exposure", "30%", f"{total_financial}", f"{int(financial_score)}", "30"),
            ("Complexity", "25%", f"{total_complexity}", f"{int(complexity_score)}", "25"),
            ("Compliance Risk", "25%", f"{total_compliance}", f"{int(compliance_score)}", "25"),
            ("Operational Risk", "20%", f"{total_operational}", f"{int(operational_score)}", "20"),
        ]

        for category, weight, raw, weighted, max_val in breakdown:
            ws.cell(row=row, column=1).value = category
            ws.cell(row=row, column=2).value = weight
            ws.cell(row=row, column=3).value = raw
            ws.cell(row=row, column=4).value = weighted
            ws.cell(row=row, column=5).value = max_val
            row += 1

        row += 2

        # Component-level risks
        ws[f'A{row}'] = "COMPONENT-LEVEL RISK SCORES"
        ws[f'A{row}'].font = self.section_font
        row += 1

        # Headers
        headers = ['Component', 'Financial', 'Complexity', 'Compliance', 'Operational']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        row += 1

        # Components
        for comp_id in sorted(plan.get_all_component_ids()):
            component = self.library.get_component(comp_id)
            ws.cell(row=row, column=1).value = f"{component.component_id}: {component.name}"
            ws.cell(row=row, column=2).value = component.governance_impact.financial_risk
            ws.cell(row=row, column=3).value = component.governance_impact.complexity
            ws.cell(row=row, column=4).value = component.governance_impact.compliance_risk
            ws.cell(row=row, column=5).value = component.governance_impact.operational_risk
            row += 1

        # Format columns
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
