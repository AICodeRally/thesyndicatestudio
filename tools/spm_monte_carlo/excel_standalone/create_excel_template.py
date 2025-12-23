"""
Create Excel template with all sheets, formatting, and sample data.
This creates a complete SPM_Monte_Carlo.xlsm file ready to use.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from pathlib import Path


def create_excel_template():
    """Create complete Excel template."""

    print("=" * 70)
    print("CREATING SPM MONTE CARLO EXCEL TEMPLATE")
    print("=" * 70)

    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=16, color="366092")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 1. Create Dashboard Sheet
    print("\n1. Creating Dashboard sheet...")
    ws_dashboard = wb.create_sheet("Dashboard", 0)

    # Title
    ws_dashboard['A1'] = "SPM MONTE CARLO SIMULATION TOOL"
    ws_dashboard['A1'].font = Font(bold=True, size=18, color="366092")
    ws_dashboard.merge_cells('A1:F1')

    # Subtitle
    ws_dashboard['A2'] = "Sales Performance Management - Compensation Analysis"
    ws_dashboard['A2'].font = Font(size=11, italic=True)
    ws_dashboard.merge_cells('A2:F2')

    # Instructions section
    ws_dashboard['A4'] = "INSTRUCTIONS:"
    ws_dashboard['A4'].font = Font(bold=True, size=12)

    instructions = [
        "1. Add your historical performance data to the 'Historical_Data' sheet",
        "2. Configure your compensation plan in the 'Plan_Config' sheet",
        "3. Adjust simulation settings in the 'Config' sheet (optional)",
        "4. Click the 'Run Simulation' button below",
        "5. View results in the 'Results' sheet when complete"
    ]

    for i, instr in enumerate(instructions, start=5):
        ws_dashboard[f'A{i}'] = instr
        ws_dashboard[f'A{i}'].font = Font(size=10)

    # Status section
    ws_dashboard['A11'] = "Current Status:"
    ws_dashboard['A11'].font = Font(bold=True)
    ws_dashboard['B11'] = "Ready"
    ws_dashboard['B11'].font = Font(color="00AA00")

    # Button placeholder (VBA will add actual button)
    ws_dashboard['A13'] = "▶ RUN SIMULATION"
    ws_dashboard['A13'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_dashboard['A13'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_dashboard['A13'].alignment = Alignment(horizontal='center', vertical='center')
    ws_dashboard.merge_cells('A13:B13')
    ws_dashboard.row_dimensions[13].height = 30

    # Notes
    ws_dashboard['A15'] = "NOTES:"
    ws_dashboard['A15'].font = Font(bold=True)
    ws_dashboard['A16'] = "• Simulation typically takes 30-120 seconds depending on data size"
    ws_dashboard['A17'] = "• A command window will show progress during simulation"
    ws_dashboard['A18'] = "• Do not close Excel while simulation is running"

    # Format columns
    ws_dashboard.column_dimensions['A'].width = 60
    ws_dashboard.column_dimensions['B'].width = 20

    # 2. Create Historical_Data Sheet
    print("2. Creating Historical_Data sheet...")
    ws_hist = wb.create_sheet("Historical_Data", 1)

    # Headers
    headers = ['rep_id', 'period', 'quota', 'actual_sales', 'quota_attainment', 'deal_count', 'avg_deal_size']
    for col, header in enumerate(headers, start=1):
        cell = ws_hist.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Sample data
    sample_data = [
        ['REP001', '2024-01', 100000, 95000, 0.95, 12, 7916.67],
        ['REP001', '2024-02', 100000, 108000, 1.08, 15, 7200.00],
        ['REP001', '2024-03', 100000, 87000, 0.87, 10, 8700.00],
        ['REP002', '2024-01', 75000, 82000, 1.09, 20, 4100.00],
        ['REP002', '2024-02', 75000, 71000, 0.95, 18, 3944.44],
        ['REP002', '2024-03', 75000, 79000, 1.05, 19, 4157.89],
    ]

    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_hist.cell(row=row_idx, column=col_idx).value = value

    # Format columns
    ws_hist.column_dimensions['A'].width = 12
    ws_hist.column_dimensions['B'].width = 12
    ws_hist.column_dimensions['C'].width = 15
    ws_hist.column_dimensions['D'].width = 15
    ws_hist.column_dimensions['E'].width = 18
    ws_hist.column_dimensions['F'].width = 12
    ws_hist.column_dimensions['G'].width = 15

    # Add instructions at top
    ws_hist.insert_rows(1)
    ws_hist['A1'] = "📝 INSTRUCTIONS: Replace sample data below with your historical performance data. Keep column headers exactly as shown."
    ws_hist['A1'].font = Font(italic=True, color="0000FF")
    ws_hist.merge_cells('A1:G1')

    # 3. Create Plan_Config Sheet
    print("3. Creating Plan_Config sheet...")
    ws_plan = wb.create_sheet("Plan_Config", 2)

    # Title
    ws_plan['A1'] = "COMPENSATION PLAN CONFIGURATION"
    ws_plan['A1'].font = title_font
    ws_plan.merge_cells('A1:D1')

    # Plan details
    ws_plan['A3'] = "Plan ID:"
    ws_plan['A3'].font = Font(bold=True)
    ws_plan['B3'] = "PLAN_2024_SALES"

    ws_plan['A4'] = "Plan Name:"
    ws_plan['A4'].font = Font(bold=True)
    ws_plan['B4'] = "2024 Sales Commission Plan"

    # Commission Tiers section
    ws_plan['A6'] = "COMMISSION TIERS"
    ws_plan['A6'].font = Font(bold=True, size=12)
    ws_plan.merge_cells('A6:D6')

    tier_headers = ['tier_name', 'quota_min', 'quota_max', 'rate']
    for col, header in enumerate(tier_headers, start=1):
        cell = ws_plan.cell(row=7, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Sample tiers
    tiers = [
        ['Tier 1', 0.00, 0.75, 0.02],
        ['Tier 2', 0.75, 1.00, 0.03],
        ['Tier 3 (Accelerator)', 1.00, 1.25, 0.045],
        ['Tier 4 (Super Accelerator)', 1.25, 10.00, 0.06],
    ]

    for row_idx, tier in enumerate(tiers, start=8):
        for col_idx, value in enumerate(tier, start=1):
            ws_plan.cell(row=row_idx, column=col_idx).value = value

    # Bonuses section
    ws_plan['A13'] = "BONUSES (Optional)"
    ws_plan['A13'].font = Font(bold=True, size=12)
    ws_plan.merge_cells('A13:F13')

    bonus_headers = ['bonus_name', 'trigger_metric', 'trigger_condition', 'trigger_value', 'payout', 'frequency']
    for col, header in enumerate(bonus_headers, start=1):
        cell = ws_plan.cell(row=14, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Sample bonus
    bonus = ['100% Club Bonus', 'quota_attainment', '>=', 1.0, 5000, 'quarterly']
    for col_idx, value in enumerate(bonus, start=1):
        ws_plan.cell(row=15, column=col_idx).value = value

    # Format columns
    ws_plan.column_dimensions['A'].width = 25
    ws_plan.column_dimensions['B'].width = 15
    ws_plan.column_dimensions['C'].width = 15
    ws_plan.column_dimensions['D'].width = 12
    ws_plan.column_dimensions['E'].width = 12
    ws_plan.column_dimensions['F'].width = 12

    # 4. Create Config Sheet
    print("4. Creating Config sheet...")
    ws_config = wb.create_sheet("Config", 3)

    ws_config['A1'] = "SIMULATION CONFIGURATION"
    ws_config['A1'].font = title_font
    ws_config.merge_cells('A1:C1')

    config_items = [
        ['Iterations:', 10000, 'Number of simulation runs (1000-100000)'],
        ['Random Seed:', 42, 'For reproducible results (any integer)'],
        ['Sampling Strategy:', 'monte_carlo', 'Options: monte_carlo, lhs, quasi_random'],
    ]

    for row_idx, (label, value, note) in enumerate(config_items, start=3):
        ws_config.cell(row=row_idx, column=1).value = label
        ws_config.cell(row=row_idx, column=1).font = Font(bold=True)
        ws_config.cell(row=row_idx, column=2).value = value
        ws_config.cell(row=row_idx, column=3).value = note
        ws_config.cell(row=row_idx, column=3).font = Font(italic=True, size=9)

    ws_config.column_dimensions['A'].width = 20
    ws_config.column_dimensions['B'].width = 15
    ws_config.column_dimensions['C'].width = 45

    # 5. Create Results Sheet
    print("5. Creating Results sheet...")
    ws_results = wb.create_sheet("Results", 4)

    ws_results['A1'] = "SIMULATION RESULTS"
    ws_results['A1'].font = title_font
    ws_results['A2'] = "(Results will appear here after running simulation)"
    ws_results['A2'].font = Font(italic=True, color="666666")

    ws_results.column_dimensions['A'].width = 30
    ws_results.column_dimensions['B'].width = 20

    # Save as .xlsx first
    output_dir = Path(__file__).parent / 'dist' / 'SPM_Monte_Carlo'
    output_dir.mkdir(parents=True, exist_ok=True)

    xlsx_path = output_dir / 'SPM_Monte_Carlo.xlsx'

    print(f"\n6. Saving workbook to {xlsx_path}...")
    wb.save(xlsx_path)

    print("\n" + "=" * 70)
    print("✅ EXCEL TEMPLATE CREATED SUCCESSFULLY!")
    print("=" * 70)
    print(f"\nFile location: {xlsx_path}")
    print(f"File size: {xlsx_path.stat().st_size / 1024:.1f} KB")

    print("\nNEXT STEPS:")
    print("1. Open the Excel file")
    print("2. Save As → Excel Macro-Enabled Workbook (.xlsm)")
    print("3. Press Alt+F11 to open VBA editor")
    print("4. Insert > Module")
    print("5. Copy code from templates/VBA_Code.txt")
    print("6. Save and test!")

    print("\nOR:")
    print("I can add the VBA code programmatically using python-docx2txt...")

    return xlsx_path


if __name__ == '__main__':
    try:
        create_excel_template()
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
